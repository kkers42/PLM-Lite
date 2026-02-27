"""
PLM Lite V1.0 — Excel BOM export
"""
from __future__ import annotations

import io
from typing import Any

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


def _header_style():
    return {
        "font": Font(bold=True, color="FFFFFF"),
        "fill": PatternFill("solid", fgColor="1A73E8"),
        "alignment": Alignment(horizontal="center"),
        "border": Border(
            bottom=Side(style="thin", color="FFFFFF"),
        ),
    }


def _apply(cell: Any, **styles) -> None:
    for attr, val in styles.items():
        setattr(cell, attr, val)


def generate_bom_excel(
    root_part: dict,
    bom_rows: list[dict],
) -> bytes:
    """
    Generate an Excel BOM workbook and return bytes.
    root_part: dict with part_number, part_name, part_revision
    bom_rows:  list from db.get_bom_flat() — depth, quantity, part_number, etc.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "BOM"

    # Title row
    ws.merge_cells("A1:H1")
    title_cell = ws["A1"]
    title_cell.value = (
        f"Bill of Materials — {root_part['part_number']} Rev {root_part['part_revision']} — {root_part['part_name']}"
    )
    title_cell.font = Font(bold=True, size=13)
    title_cell.alignment = Alignment(horizontal="left")
    ws.row_dimensions[1].height = 22

    # Header row
    headers = ["Level", "Part Number", "Part Name", "Revision", "Quantity", "Type", "Release Status", "Notes"]
    col_widths = [8, 18, 32, 10, 10, 14, 16, 30]

    hstyle = _header_style()
    for col, (header, width) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=2, column=col, value=header)
        _apply(cell, **hstyle)
        ws.column_dimensions[cell.column_letter].width = width

    # Data rows
    alt_fill = PatternFill("solid", fgColor="EAF2FB")
    released_font = Font(color="188038")

    for i, row in enumerate(bom_rows, start=3):
        depth = row.get("depth", 0)
        indent = "  " * depth
        is_even = i % 2 == 0

        values = [
            depth,
            indent + row.get("part_number", ""),
            row.get("part_name", ""),
            row.get("part_revision", ""),
            row.get("quantity", 1.0),
            row.get("relationship_type", "assembly"),
            row.get("release_status", ""),
            "",
        ]

        for col, val in enumerate(values, start=1):
            cell = ws.cell(row=i, column=col, value=val)
            if is_even:
                cell.fill = alt_fill
            if row.get("release_status") == "Released":
                cell.font = released_font

    # Freeze header rows
    ws.freeze_panes = "A3"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
